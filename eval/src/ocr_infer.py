#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
古籍OCR批量推理脚本（SongPanda-Bench）

支持两种后端：
  1) 兼容 OpenAI 协议的远程 API（例如阿里云 DashScope、OpenRouter、本地 vLLM 等）
  2) 基于 PaddleFormers 的本地推理

所有配置均通过命令行参数或环境变量传入，代码中不硬编码任何 API Key。

用法示例
-----------------------------------------------------------------
# 方式一：调用兼容 OpenAI 协议的 API
export OPENAI_API_KEY=sk-xxxxxxxx
python ocr_infer.py \
    --image-folder ../benchmark \
    --output ./songpanda_pred.xlsx \
    --backend openai \
    --base-url https://dashscope.aliyuncs.com/compatible-mode/v1 \
    --model qwen2.5-vl-7b-instruct

# 方式二：本地 PaddleFormers 推理
python ocr_infer.py \
    --image-folder ../benchmark \
    --output ./songpanda_pred.xlsx \
    --backend hf \
    --model-path /path/to/SongPanda2.0
"""

import os
import re
import sys
import time
import base64
import argparse
import threading
from pathlib import Path
from datetime import datetime

import pandas as pd
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


# =========================================================================
# 默认 Prompt（与 SongPanda 训练一致）
# =========================================================================
DEFAULT_PROMPT = (
    " OCR:<footnote><head>"
)


# =========================================================================
# 后端：兼容 OpenAI 协议的远程 API
# =========================================================================
class OpenAIBackend:
    def __init__(self, model: str, base_url: str, api_key: str,
                 max_tokens: int = 2048, temperature: float = 0.1):
        from openai import OpenAI
        if not api_key:
            raise ValueError(
                "未提供 API Key。请通过 --api-key 参数或环境变量 OPENAI_API_KEY 传入。"
            )
        self.client = OpenAI(api_key=api_key, base_url=base_url)
        self.model = model
        self.max_tokens = max_tokens
        self.temperature = temperature

    @staticmethod
    def _encode_image(image_path: str) -> str:
        with open(image_path, "rb") as f:
            return base64.b64encode(f.read()).decode("utf-8")

    def infer(self, image_path: str, prompt: str) -> str:
        b64 = self._encode_image(image_path)
        messages = [{
            "role": "user",
            "content": [
                {"type": "text", "text": prompt},
                {"type": "image_url",
                 "image_url": {"url": f"data:image/jpeg;base64,{b64}"}},
            ],
        }]
        resp = self.client.chat.completions.create(
            model=self.model,
            messages=messages,
            max_tokens=self.max_tokens,
            temperature=self.temperature,
        )
        return resp.choices[0].message.content


# =========================================================================
# 后端：基于 PaddleFormers 的本地推理
# =========================================================================
class HFBackend:
    def __init__(self, model_path: str, max_new_tokens: int = 2048,
                 temperature: float = 0.1):
        import paddle
        from paddleformers.transformers import AutoModelForConditionalGeneration, AutoProcessor
        from paddleformers.generation import GenerationConfig
        self.paddle = paddle
        self.processor = AutoProcessor.from_pretrained(
            model_path, trust_remote_code=True
        )
        self.model = AutoModelForConditionalGeneration.from_pretrained(
            model_path,
            convert_from_hf=True,
        ).eval()
        self.model.config._attn_implementation = "flashmask"
        self.model.visual.config._attn_implementation = "flashmask"
        self.max_new_tokens = max_new_tokens
        self.temperature = temperature
        self.gen_config = GenerationConfig(
            do_sample=temperature > 0,
            temperature=max(temperature, 1e-5),
            bos_token_id=1,
            eos_token_id=2,
            pad_token_id=0,
            use_cache=True,
        )

    def infer(self, image_path: str, prompt: str) -> str:
        from PIL import Image
        image = Image.open(image_path).convert("RGB")
        messages = [{
            "role": "user",
            "content": [
                {"type": "image", "image": image},
                {"type": "text", "text": prompt},
            ],
        }]
        inputs = self.processor.apply_chat_template(
            messages,
            tokenize=True,
            add_generation_prompt=True,
            return_dict=True,
            return_tensors="pd",
        )
        with self.paddle.no_grad():
            output_ids = self.model.generate(
                **inputs,
                generation_config=self.gen_config,
                max_new_tokens=self.max_new_tokens,
            )
        text = self.processor.decode(output_ids[0].tolist()[0], skip_special_tokens=True)
        return text.strip()


# =========================================================================
# 鲁棒性判定：是否包含连续超过 5 个重复的 n-gram 片段
# =========================================================================
def has_excessive_repetition(text: str, n: int = 5, max_repeat: int = 5) -> bool:
    """
    与论文中"鲁棒性R"的判定规则一致：
    模型输出如果存在连续超过 max_repeat(=5) 个重复的 n-gram 片段，
    标志为不通过，需要重试。
    """
    if not text or len(text) < n * (max_repeat + 1):
        return False
    for i in range(len(text) - n * (max_repeat + 1) + 1):
        gram = text[i:i + n]
        ok = True
        for k in range(1, max_repeat + 1):
            if text[i + k * n:i + (k + 1) * n] != gram:
                ok = False
                break
        if ok:
            return True
    return False


# =========================================================================
# Excel 实时写入
# =========================================================================
class RealTimeExcelWriter:
    """实时 Excel 写入器，列结构与评测脚本 evaluate_ocr.py 对齐。"""

    COLUMNS = ['序号', '图片名称', '图片路径', '处理状态',
               '重试次数', '处理时间(秒)', 'OCR结果', '处理时间戳']

    def __init__(self, filename: str, resume_mode: bool = False):
        self.filename = filename
        self.lock = threading.Lock()
        self.data = []
        if resume_mode and os.path.exists(filename):
            try:
                df = pd.read_excel(filename)
                self.data = df.to_dict('records')
                print(f"📁 续写模式：继续写入 {os.path.abspath(filename)}")
            except Exception as e:
                print(f"⚠️  读取 Excel 失败，创建新文件: {e}")
                self._create_new_excel()
        else:
            self._create_new_excel()

    def _create_new_excel(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "古籍OCR识别结果"
        ws.append(self.COLUMNS)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092",
                                  end_color="366092", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        for col in range(1, len(self.COLUMNS) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        widths = [8, 25, 50, 12, 12, 15, 80, 20]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
        os.makedirs(os.path.dirname(os.path.abspath(filename)) or ".", exist_ok=True)
        wb.save(self.filename)
        print(f"📁 新建 Excel 文件: {os.path.abspath(self.filename)}")

    def write(self, image_path: str, result: str, elapsed: float,
              status: str, retry_count: int):
        with self.lock:
            idx = None
            for i, row in enumerate(self.data):
                if row.get('图片路径') == image_path:
                    idx = i
                    break
            row = {
                '序号': len(self.data) + 1 if idx is None else idx + 1,
                '图片名称': os.path.basename(image_path),
                '图片路径': image_path,
                '处理状态': status,
                '重试次数': retry_count,
                '处理时间(秒)': round(elapsed, 2),
                'OCR结果': result,
                '处理时间戳': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            }
            if idx is None:
                self.data.append(row)
            else:
                self.data[idx] = row
            self._flush()

    def _flush(self):
        try:
            df = pd.DataFrame(self.data)
            df['序号'] = range(1, len(df) + 1)
            with pd.ExcelWriter(self.filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='古籍OCR识别结果', index=False)
        except Exception as e:
            print(f"⚠️  保存 Excel 失败: {e}")


# =========================================================================
# 文件发现 & 排序
# =========================================================================
IMAGE_EXTS = ['.jpg', '.jpeg', '.png', '.bmp', '.tiff',
              '.JPG', '.JPEG', '.PNG']


def page_sort_key(filename: str):
    """按页码排序，同页内 top 先于 bottom。"""
    name = os.path.basename(filename).lower()
    nums = re.findall(r'(\d+)', name)
    page = int(nums[0]) if nums else float('inf')
    m = re.search(r'[_-](top|bottom)(?=\.)', name)
    if m:
        variant_order = 0 if m.group(1) == 'top' else 1
    else:
        variant_order = 2
    return (page, variant_order, name)


def find_image_files(image_folder: str):
    """递归一层，返回排序后的图片绝对路径列表。"""
    root = Path(image_folder)
    if not root.exists():
        print(f"❌ 图片文件夹不存在: {image_folder}")
        return []
    files = []
    for ext in IMAGE_EXTS:
        for p in root.glob(f"*{ext}"):
            if not p.name.startswith('.'):
                files.append(str(p.absolute()))
    for sub in root.iterdir():
        if sub.is_dir() and not sub.name.startswith('.'):
            for ext in IMAGE_EXTS:
                for p in sub.glob(f"*{ext}"):
                    if not p.name.startswith('.'):
                        files.append(str(p.absolute()))
    files.sort(key=page_sort_key)
    return files


def parse_processed(xlsx_path: str):
    """从已有结果文件中解析已处理图片，区分成功与失败。"""
    success, failed = set(), set()
    if not os.path.exists(xlsx_path):
        return success, failed
    try:
        df = pd.read_excel(xlsx_path, sheet_name='古籍OCR识别结果')
        if '图片路径' in df.columns and '处理状态' in df.columns:
            for _, row in df.iterrows():
                p = str(row['图片路径']) if pd.notna(row['图片路径']) else None
                s = str(row['处理状态']) if pd.notna(row['处理状态']) else None
                if p:
                    (success if s == "成功" else failed).add(p)
        print(f"📋 续传：成功 {len(success)} / 失败 {len(failed)}")
    except Exception as e:
        print(f"❌ 解析续传文件失败: {e}")
    return success, failed


# =========================================================================
# 单张图片推理（含鲁棒性重试）
# =========================================================================
def infer_with_retry(backend, image_path: str, prompt: str,
                     max_retries: int, retry_interval: float,
                     ngram_n: int, ngram_max_repeat: int):
    """
    对单张图片进行推理，如果输出包含连续超过 ngram_max_repeat 个重复的
    ngram_n-gram，则视为失败并重试，最多 max_retries 次。

    返回：(result, elapsed, status, retry_count)
    retry_count：实际重试次数（首次成功 = 0）
    """
    last_result = ""
    last_elapsed = 0.0
    for attempt in range(max_retries + 1):
        t0 = time.time()
        try:
            text = backend.infer(image_path, prompt)
            elapsed = time.time() - t0
            last_result, last_elapsed = text, elapsed
            if has_excessive_repetition(text, n=ngram_n,
                                        max_repeat=ngram_max_repeat):
                tqdm.write(f"🔁 第{attempt + 1}次 {os.path.basename(image_path)}: "
                           f"检测到重复 n-gram，准备重试")
                if attempt < max_retries:
                    time.sleep(retry_interval)
                    continue
                return text, elapsed, "失败", attempt
            return text, elapsed, "成功", attempt
        except Exception as e:
            elapsed = time.time() - t0
            tqdm.write(f"❌ 第{attempt + 1}次 {os.path.basename(image_path)}: {e}")
            last_result, last_elapsed = f"API调用失败: {e}", elapsed
            if attempt < max_retries:
                time.sleep(retry_interval)
    return last_result, last_elapsed, "失败", max_retries


# =========================================================================
# 批量处理
# =========================================================================
def batch_run(args, backend):
    image_files = find_image_files(args.image_folder)
    if not image_files:
        print("❌ 未找到任何图片")
        return

    successful, failed = set(), set()
    resume_mode = bool(args.resume)
    output = args.resume if args.resume else args.output

    if resume_mode:
        successful, failed = parse_processed(output)
        image_files = [f for f in image_files if f not in successful]

    if not image_files:
        print("✅ 所有图片已处理完成")
        return

    print(f"🚀 待处理图片数: {len(image_files)}")
    print(f"💾 输出文件: {output}")
    print("=" * 60)

    writer = RealTimeExcelWriter(output, resume_mode=resume_mode)
    success_cnt = fail_cnt = 0
    total_time = 0.0

    with tqdm(total=len(image_files), ncols=100,
              bar_format="{l_bar}{bar}| {n_fmt}/{total_fmt} "
                         "[{elapsed}<{remaining}, {rate_fmt}]") as pbar:
        for img in image_files:
            pbar.set_description(f"OCR: {os.path.basename(img)[:15]:<15}")
            result, elapsed, status, retry_cnt = infer_with_retry(
                backend, img, args.prompt,
                max_retries=args.max_retries,
                retry_interval=args.retry_interval,
                ngram_n=args.ngram_n,
                ngram_max_repeat=args.ngram_max_repeat,
            )
            if status == "成功":
                success_cnt += 1
                total_time += elapsed
            else:
                fail_cnt += 1
            writer.write(img, result, elapsed, status, retry_cnt)
            pbar.set_postfix({
                '成功': success_cnt,
                '失败': fail_cnt,
                '平均耗时': f"{(total_time / success_cnt):.1f}s"
                            if success_cnt else "-",
            })
            pbar.update(1)

    print("\n🎉 处理完成")
    print(f"   ✅ 成功 {success_cnt}  ❌ 失败 {fail_cnt}")
    if success_cnt:
        print(f"   ⏱️  平均耗时 {total_time / success_cnt:.2f} 秒")
    print(f"   📂 结果: {os.path.abspath(output)}")


# =========================================================================
# CLI
# =========================================================================
def parse_args():
    parser = argparse.ArgumentParser(
        description="SongPanda 古籍OCR批量推理脚本（无硬编码）",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument('--image-folder', required=True,
                        help='古籍图像文件夹（会递归搜索一层子目录）')
    parser.add_argument('--output', default=None,
                        help='输出 Excel 路径（默认：ocr_results_<时间戳>.xlsx）')
    parser.add_argument('--resume', default=None,
                        help='续传 Excel 路径；指定后将跳过已成功的样本')

    parser.add_argument('--backend', choices=['openai', 'hf'], default='openai',
                        help='推理后端：openai 兼容 API 或本地 HuggingFace')

    # OpenAI 兼容后端
    parser.add_argument('--base-url', default=os.getenv('OPENAI_BASE_URL',
                        'https://dashscope.aliyuncs.com/compatible-mode/v1'),
                        help='OpenAI 兼容 API base_url（也可用环境变量 OPENAI_BASE_URL）')
    parser.add_argument('--api-key', default=os.getenv('OPENAI_API_KEY', ''),
                        help='API Key（也可用环境变量 OPENAI_API_KEY；'
                             '请勿写死在代码里）')
    parser.add_argument('--model', default='qwen2.5-vl-7b-instruct',
                        help='OpenAI 兼容后端下的模型名')

    # HF 本地后端
    parser.add_argument('--model-path', default=None,
                        help='本地推理时的模型路径或仓库名')

    # 推理控制
    parser.add_argument('--prompt', default=DEFAULT_PROMPT, help='OCR 指令')
    parser.add_argument('--max-tokens', type=int, default=2048)
    parser.add_argument('--temperature', type=float, default=0.1)

    # 鲁棒性 / 重试（对齐论文：5-gram 重复超过 5 次视为失败，最多重试 20 次）
    parser.add_argument('--max-retries', type=int, default=20,
                        help='最大重试次数（鲁棒性上限）')
    parser.add_argument('--retry-interval', type=float, default=2.0,
                        help='每次重试间隔（秒）')
    parser.add_argument('--ngram-n', type=int, default=5,
                        help='鲁棒性判定用的 n-gram 长度')
    parser.add_argument('--ngram-max-repeat', type=int, default=5,
                        help='允许的连续重复次数上限')

    return parser.parse_args()


def main():
    args = parse_args()
    if not os.path.exists(args.image_folder):
        print(f"❌ 图片文件夹不存在: {args.image_folder}")
        sys.exit(1)

    if not args.output and not args.resume:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        args.output = f"ocr_results_{ts}.xlsx"

    if args.backend == 'openai':
        backend = OpenAIBackend(
            model=args.model,
            base_url=args.base_url,
            api_key=args.api_key,
            max_tokens=args.max_tokens,
            temperature=args.temperature,
        )
    else:  # hf
        if not args.model_path:
            print("❌ --backend hf 时必须指定 --model-path")
            sys.exit(1)
        backend = HFBackend(
            model_path=args.model_path,
            max_new_tokens=args.max_tokens,
            temperature=args.temperature,
        )

    print("🚀 SongPanda OCR 批量推理")
    print("=" * 60)
    print(f"📁 图片文件夹: {args.image_folder}")
    print(f"⚙️  后端: {args.backend}")
    print(f"🤖 模型: "
          f"{args.model_path if args.backend == 'hf' else args.model}")
    print("=" * 60)

    batch_run(args, backend)


if __name__ == "__main__":
    main()
